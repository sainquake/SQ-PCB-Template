import json
import os

def readData(name):
    if not os.path.isfile(name): 
        print("File path {} does not exist...".format(name))
        return None
    else:
        with open(name, 'r', encoding='utf8') as f: 
            j = json.load(f)
            f.close()
        return j
def writeData(name,data):
    if data!=None:
        if (len(name.split("/"))>1) and (not os.path.isdir(name.split("/")[0])):
            os.mkdir(name.split("/")[0])
        with open(name, 'w', encoding='utf8') as outfile: 
            json.dump(data, outfile, indent=4, ensure_ascii=False)
            outfile.close()
    else:
        print(f'Data for {name} is none ')    

paths = readData('paths.json')


#EXTRACT Rv1 and Rv2 from BOM
BOMPath = ''
BOM = ''
for item in paths:
    if 'BOM' in item['key']:
        print(item['name'])
        BOMPath = item['path'] +'/'
        BOM = item['path'] +'/'+item['name']
        break

print (BOM)
if not os.path.isfile(BOM):
    print('BOM NOT FOUND')
    raise Exception(f'BOM NOT FOUND')

BOMList = []
try:
    from openpyxl import Workbook
    from openpyxl import load_workbook

    wb_obj = Workbook()
    wb_obj.template = True
    wb_obj = load_workbook(BOM) 
    wb_obj.template = True
    #
    sheet = wb_obj.active

    ns = []
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        #print(i,row)
        if i==0:
            ns = row
        else:
            li = {}
            for key,value in enumerate(row):
                #print(key,value)
                li[ns[key]] = value
            #print(li)
            BOMList.append(li)
except:
    print('TRY BOM FROM TXT')

    if not os.path.isfile(BOMPath+'BOMtxt-BOM.txt'):
        print('BOM in TXT format NOT FOUND')
        raise Exception(f'BOM in TXT format NOT FOUND')

    bom = open(BOMPath+'BOMtxt-BOM.txt',mode="r",encoding="ISO-8859-1")
    i=0
    ns = []
    while bom:
        line = bom.readline()
        if i==0:
            ns = line.split('\t')
        else:
            if len(line)>1:
                row = line.split('\t')
                li = {}
                print(i,len(line),line.split('\t')[0])
                for key,value in enumerate(row):
                    li[ns[key]] = value.replace('"','')
                BOMList.append(li)
        i+=1
        if len(line)<1:
            break
    bom.close()

writeData('BOMList.json',BOMList)